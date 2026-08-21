"""Tests for agents/parser.py — Agent 1.

Fixtures with formulas are recalculated through LibreOffice before the parser
reads them. That step is load-bearing, not hygiene: openpyxl cannot calculate,
so without it every cached_value would be None and the assertions that matter
here would pass while proving nothing.
"""

import hashlib
import pathlib
import re

import openpyxl
import pytest
from openpyxl.workbook.properties import CalcProperties

from agents.parser import parse_workbook
from core.workbook_identity import sha256_bytes
from fixture_helpers import (
    libreoffice_available,
    recalculate_workbook,
    set_calc_mode,
    strip_calc_pr,
)


def _bytes(path) -> bytes:
    """Read a fixture workbook as bytes.

    parse_workbook takes bytes, not a path: the bytes confirmed at Gate 1 must
    be the exact bytes parsed, and a path cannot carry that guarantee across the
    parser's five separate reads.
    """
    return pathlib.Path(path).read_bytes()


needs_libreoffice = pytest.mark.skipif(
    not libreoffice_available(),
    reason="LibreOffice is required to give fixtures real cached formula values",
)


def _clean_workbook(path):
    wb = openpyxl.Workbook()
    inputs = wb.active
    inputs.title = "Inputs"
    inputs["A1"] = 100
    inputs["A2"] = 200

    provisions = wb.create_sheet("Provisions")
    provisions["C1"] = 10
    provisions["C2"] = 20
    provisions["C3"] = 30
    provisions["C4"] = 40
    provisions["C5"] = "=SUM(C1:C4)"
    provisions["D1"] = "=Inputs!A1*2"
    provisions["E1"] = "=C5+D1"
    provisions["C5"].number_format = "#,##0.00"

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Test 1 — a clean, genuinely calculated workbook
# ---------------------------------------------------------------------------


@needs_libreoffice
def test_clean_workbook_is_parsed_with_real_cached_values(tmp_path):
    path = str(tmp_path / "clean.xlsx")
    _clean_workbook(path)
    # LOAD-BEARING: without this line every cached_value below is None and the
    # formula-and-value assertion becomes vacuous. Do not remove it to make a
    # test pass — if it fails, the fixture is wrong, not the assertion.
    recalculate_workbook(path)

    parsed = parse_workbook(_bytes(path))

    assert parsed.tab_names == ["Inputs", "Provisions"]

    # The property CellRecord exists for: one record, both facts.
    total = parsed.cells["Provisions!C5"]
    assert total.formula is not None and total.formula.startswith("=SUM")
    assert total.cached_value == 100
    assert total.is_stale is False

    both_populated = [
        c for c in parsed.cells.values() if c.formula is not None and c.cached_value is not None
    ]
    assert len(both_populated) >= 3


@needs_libreoffice
def test_cell_dependency_graph_has_individual_cell_edges(tmp_path):
    path = str(tmp_path / "clean.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)

    graph = parse_workbook(_bytes(path)).cell_dependency_graph

    # A range is expanded cell by cell — the graph answers "which cells feed
    # this one", which a range edge could not.
    assert set(graph["Provisions!C5"]) == {
        "Provisions!C1",
        "Provisions!C2",
        "Provisions!C3",
        "Provisions!C4",
    }
    # Cross-tab references resolve to the other tab's cell, not just the tab.
    assert graph["Provisions!D1"] == ["Inputs!A1"]
    assert set(graph["Provisions!E1"]) == {"Provisions!C5", "Provisions!D1"}


@needs_libreoffice
def test_tab_graph_is_coarse_and_separate_from_the_cell_graph(tmp_path):
    path = str(tmp_path / "clean.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)

    parsed = parse_workbook(_bytes(path))
    assert parsed.tab_dependency_graph["Provisions"] == ["Inputs"]
    assert "Provisions" not in parsed.cell_dependency_graph
    assert "Provisions!C5" not in parsed.tab_dependency_graph


@needs_libreoffice
def test_workbook_hash_is_sha256_of_the_file_and_changes_with_it(tmp_path):
    path = str(tmp_path / "clean.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)

    first = parse_workbook(_bytes(path)).workbook_meta.workbook_hash
    assert re.fullmatch(r"[0-9a-f]{64}", first)
    with open(path, "rb") as handle:
        assert first == hashlib.sha256(handle.read()).hexdigest()

    wb = openpyxl.load_workbook(path)
    wb["Inputs"]["A1"] = 999
    wb.save(path)
    assert parse_workbook(_bytes(path)).workbook_meta.workbook_hash != first


@needs_libreoffice
def test_number_format_is_captured(tmp_path):
    path = str(tmp_path / "clean.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)
    assert parse_workbook(_bytes(path)).cells["Provisions!C5"].number_format == "#,##0.00"


# ---------------------------------------------------------------------------
# Test 2 — a deliberately messy workbook
# ---------------------------------------------------------------------------


def _messy_workbook(path):
    """Deliberately NOT recalculated.

    Every formula cell here therefore has no cached value, which is the exact
    condition is_stale is meant to catch. Recalculating this fixture would
    destroy the thing it exists to test.
    """
    wb = openpyxl.Workbook()

    provisions = wb.active
    provisions.title = "Provisions"
    provisions["B1"] = "#REF!"
    provisions["B2"] = "1,234"
    provisions["B3"] = 42
    provisions["B4"] = "=B3*2"
    provisions.merge_cells("D1:E1")

    # Trailing space and different case: two distinct tabs to Excel, near
    # certainly a mistake to a human.
    lookalike = wb.create_sheet("provisions ")
    lookalike["A1"] = 1

    wb.create_sheet("Blank")

    wb.save(path)
    return path


def test_messy_workbook_parses_without_raising(tmp_path):
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    parsed = parse_workbook(_bytes(path))
    assert parsed.tab_names == ["Provisions", "provisions ", "Blank"]


def test_blank_tab_is_recorded_with_zero_cells(tmp_path):
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    parsed = parse_workbook(_bytes(path))

    assert "Blank" in parsed.tab_names
    assert [k for k in parsed.cells if k.startswith("Blank!")] == []


def test_cached_error_value_is_classified_not_swallowed(tmp_path):
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    cell = parse_workbook(_bytes(path)).cells["Provisions!B1"]

    assert cell.data_type == "error"
    assert cell.is_error is True
    assert cell.error_type == "#REF!"


def test_number_stored_as_text_stays_text_and_earns_a_warning(tmp_path):
    """The parser reports what it found. Deciding "1,234" was really a number
    would silently repair a defect in the source workbook."""
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    parsed = parse_workbook(_bytes(path))

    cell = parsed.cells["Provisions!B2"]
    assert cell.data_type == "text"
    assert cell.cached_value == "1,234"
    assert any("number stored as text" in w for w in parsed.warnings)


def test_never_recalculated_formula_is_stale(tmp_path):
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    parsed = parse_workbook(_bytes(path))

    cell = parsed.cells["Provisions!B4"]
    assert cell.formula == "=B3*2"
    assert cell.cached_value is None
    assert cell.is_stale is True
    assert any("never recalculated" in w for w in parsed.warnings)


def test_lookalike_tab_names_are_kept_and_warned_about(tmp_path):
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    parsed = parse_workbook(_bytes(path))

    assert "Provisions" in parsed.tab_names
    assert "provisions " in parsed.tab_names
    assert any("look like duplicates" in w for w in parsed.warnings)


def test_merged_cells_are_warned_about(tmp_path):
    path = str(tmp_path / "messy.xlsx")
    _messy_workbook(path)
    assert any("merged cell range" in w for w in parse_workbook(_bytes(path)).warnings)


def test_a_workbook_with_only_blank_tabs_does_not_crash(tmp_path):
    path = str(tmp_path / "empty.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Nothing"
    wb.save(path)

    parsed = parse_workbook(_bytes(path))
    assert parsed.tab_names == ["Nothing"]
    assert parsed.cells == {}
    assert parsed.has_vba is False


# ---------------------------------------------------------------------------
# Test 3 — calculation mode
# ---------------------------------------------------------------------------


@needs_libreoffice
def test_manual_calc_mode_makes_every_formula_cell_stale(tmp_path):
    """Recalculate first, so cached values genuinely exist, then switch the
    workbook to manual. Staleness here can only be caused by the calc mode."""
    path = str(tmp_path / "manual.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)
    set_calc_mode(path, "manual")

    parsed = parse_workbook(_bytes(path))
    assert parsed.workbook_meta.calc_mode == "manual"

    formula_cells = [c for c in parsed.cells.values() if c.formula is not None]
    assert len(formula_cells) == 3
    for cell in formula_cells:
        assert cell.cached_value is not None, "fixture lost its cached values"
        assert cell.is_stale is True

    # A literal is not made stale by the calc mode — it has nothing to recompute.
    assert parsed.cells["Provisions!C1"].is_stale is False


@needs_libreoffice
def test_automatic_calc_mode_leaves_calculated_cells_fresh(tmp_path):
    path = str(tmp_path / "auto.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)
    set_calc_mode(path, "auto")

    parsed = parse_workbook(_bytes(path))
    assert parsed.workbook_meta.calc_mode == "automatic"
    assert parsed.cells["Provisions!C5"].is_stale is False


def test_missing_calc_pr_is_unknown_never_assumed_automatic(tmp_path):
    path = str(tmp_path / "nocalcpr.xlsx")
    _clean_workbook(path)
    strip_calc_pr(path)
    assert parse_workbook(_bytes(path)).workbook_meta.calc_mode == "unknown"


@needs_libreoffice
def test_full_calc_on_load_is_read_when_present(tmp_path):
    path = str(tmp_path / "fullcalc.xlsx")
    _clean_workbook(path)
    recalculate_workbook(path)
    set_calc_mode(path, "manual", full_calc_on_load=True)
    assert parse_workbook(_bytes(path)).workbook_meta.fully_calculated_on_load is True


# ---------------------------------------------------------------------------
# named ranges, external links, VBA
# ---------------------------------------------------------------------------


def test_named_ranges_are_extracted(tmp_path):
    path = str(tmp_path / "named.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provisions"
    ws["A1"] = 5
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName("DiscountRate", attr_text="Provisions!$A$1"))
    wb.save(path)

    named = parse_workbook(_bytes(path)).named_ranges
    assert any("DiscountRate" in key for key in named)


def test_external_links_are_detected(tmp_path):
    path = str(tmp_path / "external.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provisions"
    ws["A1"] = "='[Other.xlsx]Sheet1'!A1"
    wb.save(path)

    assert parse_workbook(_bytes(path)).external_links != []


def test_no_vba_in_a_plain_xlsx(tmp_path):
    path = str(tmp_path / "plain.xlsx")
    _messy_workbook(path)
    assert parse_workbook(_bytes(path)).has_vba is False


def test_calc_properties_set_via_openpyxl_are_read(tmp_path):
    """Belt and braces: the same manual mode, set through openpyxl's API rather
    than by rewriting the zip."""
    path = str(tmp_path / "openpyxl_manual.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provisions"
    ws["A1"] = 1
    ws["A2"] = "=A1*2"
    wb.calculation = CalcProperties(calcId=191029, calcMode="manual")
    wb.save(path)

    assert parse_workbook(_bytes(path)).workbook_meta.calc_mode == "manual"


# --------------------------------------------------------------------------
# Recommendation 2 — the parser consumes the confirmed bytes, not a path
# --------------------------------------------------------------------------


def test_parsed_hash_equals_the_hash_of_the_bytes_supplied(tmp_path):
    """The invariant the Gate 1 binding rests on: what the parser reports as
    the workbook's identity is the hash of exactly the bytes it was handed."""
    path = str(tmp_path / "identity.xlsx")
    _clean_workbook(path)
    data = _bytes(path)
    assert parse_workbook(data).workbook_meta.workbook_hash == sha256_bytes(data)


def test_parser_does_not_read_the_filesystem_after_the_bytes_are_captured(tmp_path):
    """Deleting the file must not affect the parse. If it did, the parser would
    still be reading a mutable path somewhere and the invariant would be a
    claim rather than a fact."""
    path = str(tmp_path / "identity.xlsx")
    _clean_workbook(path)
    data = _bytes(path)
    pathlib.Path(path).unlink()

    parsed = parse_workbook(data)
    assert parsed.cells["Provisions!C5"].formula == "=SUM(C1:C4)"
    assert parsed.workbook_meta.workbook_hash == sha256_bytes(data)


def test_two_readers_over_the_same_bytes_agree(tmp_path):
    """Formula mode and data-only mode each get their own BytesIO over one
    immutable sequence; both must describe the same workbook."""
    path = str(tmp_path / "identity.xlsx")
    _clean_workbook(path)
    data = _bytes(path)
    first = parse_workbook(data)
    second = parse_workbook(data)
    assert first.workbook_meta.workbook_hash == second.workbook_meta.workbook_hash
    assert first.cells.keys() == second.cells.keys()


def test_changing_one_byte_changes_the_parsed_identity(tmp_path):
    """Same filename, different contents — the case Gate 1 must be able to
    distinguish."""
    path = str(tmp_path / "identity.xlsx")
    _clean_workbook(path)
    data = _bytes(path)
    assert parse_workbook(data).workbook_meta.workbook_hash != sha256_bytes(data + b"\x00")


def test_passing_a_path_instead_of_bytes_fails_loudly(tmp_path):
    """The old signature must not keep working by accident: a silently accepted
    path would reintroduce exactly the mutable reference this removed."""
    path = str(tmp_path / "identity.xlsx")
    _clean_workbook(path)
    with pytest.raises(Exception):
        parse_workbook(str(path))
