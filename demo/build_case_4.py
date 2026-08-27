"""Reproducible builder for demo Case 4 — claims reserve roll-forward.

Generates the workbook at a caller-specified path. Does not touch any
repository asset unless that path is explicitly passed in — running this
script with no arguments writes only to a temporary location.

Usage:
    python3 demo/build_case_4.py /path/to/output.xlsx

Formula catalogue used: SUM, +, -, *, /, unary minus, direct cell references.
No VLOOKUP/INDEX-MATCH, array formulas, external links, macros, or volatile
functions. Cached values are NOT written here — this script only lays out
formulas; a separate, explicitly authorised build-time recalculation step
(see demo/recalculation_provenance.json) fills in cached values before the
file is committed.
"""

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

EUR_FORMAT = '\\€#,##0;[RED]"(€"#,##0\\);\\-'

INPUT_FILL = PatternFill(start_color="FFFDE9A8", end_color="FFFDE9A8", fill_type="solid")
INPUT_FONT = Font(color="FF1F4E78")
FORMULA_FONT = Font(color="FF000000")
CROSS_SHEET_FONT = Font(color="FF2E7D32")
OUTPUT_FILL = PatternFill(start_color="FFD9F2D9", end_color="FFD9F2D9", fill_type="solid")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _title(ws: Worksheet, text: str) -> None:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT


def _build_guide(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Demo Guide"
    _title(ws, "Case 4 — Claims reserve roll-forward and GL reconciliation")
    ws["A3"] = (
        "Actuarial reserve movement, reconciled to a signed general-ledger "
        "credit balance. Synthetic demonstration data only."
    )
    ws["A5"] = "Workbook context"
    ws["B5"] = "Value"
    for label, value, row in (
        ("Entity", "Aurora General Insurance SA", 6),
        ("Period", "2025-Q4", 7),
        ("Currency", "EUR", 8),
        ("Basis", "IFRS 17 – synthetic demonstration", 9),
        ("Designated output(s)", "Controls!B4", 10),
    ):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value

    ws["A12"] = "Sign convention"
    ws["A13"] = (
        "The actuarial reserve is a positive magnitude on the Rollforward "
        "sheet. Controls!B4 converts it into a signed accounting balance: "
        "reserves are carried on the ledger as a credit, so the signed "
        "control total is negative (Controls!B4 = -Rollforward!B9)."
    )
    ws["A13"].alignment = Alignment(wrap_text=True)

    ws["A15"] = "Cell colour legend"
    ws["A16"] = "Input"
    ws["B16"] = "Blue text with pale yellow fill — editable synthetic assumption"
    ws["A17"] = "Cross-sheet formula"
    ws["B17"] = "Green text — formula linked to another worksheet"
    ws["A18"] = "Within-sheet formula"
    ws["B18"] = "Black text — formula calculated within the worksheet"
    ws["A19"] = "Designated output"
    ws["B19"] = "Pale green fill — cell selected for reconstruction and reconciliation"

    ws["A21"] = (
        "Synthetic demonstration data only. It is not derived from a "
        "policyholder, insurer, ledger, actuarial model, or production workbook."
    )
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 46


def _build_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("Inputs")
    _title(ws, "Case 4 — Inputs (EUR)")
    ws["A3"] = "All editable inputs are synthetic and clearly separated from calculations."
    ws["A5"] = "Assumption"
    ws["B5"] = "Value (EUR)"
    ws["A5"].font = HEADER_FONT
    ws["B5"].font = HEADER_FONT

    rows = [
        (2, "Opening claims reserve", 1_250_000),
        (3, "Current-period incurred claims", 480_000),
        (4, "Claims paid", 390_000),
        (5, "Assumption strengthening", 85_000),
        (6, "FX and other movement", -25_000),
    ]
    for row, label, value in rows:
        ws[f"A{row}"] = label
        cell = ws[f"B{row}"]
        cell.value = value
        cell.number_format = EUR_FORMAT
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20


def _build_rollforward(wb: Workbook) -> None:
    ws = wb.create_sheet("Rollforward")
    _title(ws, "Case 4 — Claims reserve roll-forward (EUR)")
    ws["A3"] = "Supported catalogue only: direct references, +, -, and SUM."
    ws["A5"] = "Movement"
    ws["B5"] = "EUR"
    ws["A5"].font = HEADER_FONT
    ws["B5"].font = HEADER_FONT

    entries = [
        (4, "Opening reserve", "=Inputs!B2"),
        (5, "Current-period incurred claims", "=Inputs!B3"),
        (6, "Claims paid", "=-Inputs!B4"),
        (7, "Assumption strengthening", "=Inputs!B5"),
        (8, "FX and other movement", "=Inputs!B6"),
    ]
    for row, label, formula in entries:
        ws[f"A{row}"] = label
        cell = ws[f"B{row}"]
        cell.value = formula
        cell.number_format = EUR_FORMAT
        cell.font = CROSS_SHEET_FONT

    ws["A9"] = "Closing claims reserve"
    closing = ws["B9"]
    closing.value = "=SUM(B4:B8)"
    closing.number_format = EUR_FORMAT
    closing.font = FORMULA_FONT

    ws["A11"] = "Expected closing reserve"
    expected = ws["B11"]
    expected.value = 1_400_000
    expected.number_format = EUR_FORMAT

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20


def _build_controls(wb: Workbook) -> None:
    ws = wb.create_sheet("Controls")
    _title(ws, "Case 4 — Controls and GL bridge (EUR)")
    ws["A3"] = (
        "Converts the actuarial reserve magnitude to a signed accounting "
        "credit balance for GL reconciliation. Positive actuarial reserve "
        "magnitude becomes a negative (credit) signed accounting balance."
    )
    ws["A3"].alignment = Alignment(wrap_text=True)

    # The narrative sits on A2; A3/A4/A5 carry the control-line labels
    # immediately to the left of their formulas in column B.
    ws["A2"] = ws["A3"].value
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"] = "Actuarial closing reserve"
    ws["A4"] = "Net claims reserve"
    ws["A5"] = "Movement during period"

    b3 = ws["B3"]
    b3.value = "=Rollforward!B9"
    b3.number_format = EUR_FORMAT
    b3.font = CROSS_SHEET_FONT

    b4 = ws["B4"]
    b4.value = "=-B3"
    b4.number_format = EUR_FORMAT
    b4.font = FORMULA_FONT
    b4.fill = OUTPUT_FILL
    ws["A4"].fill = OUTPUT_FILL

    b5 = ws["B5"]
    b5.value = "=Rollforward!B9-Inputs!B2"
    b5.number_format = EUR_FORMAT
    b5.font = CROSS_SHEET_FONT

    ws["A7"] = "Expected values"
    ws["A8"] = "Actuarial closing reserve"
    ws["B8"] = 1_400_000
    ws["B8"].number_format = EUR_FORMAT
    ws["A9"] = "Net claims reserve"
    ws["B9"] = -1_400_000
    ws["B9"].number_format = EUR_FORMAT
    ws["A10"] = "Movement during period"
    ws["B10"] = 150_000
    ws["B10"].number_format = EUR_FORMAT

    ws["A12"] = "Cell colour legend"
    ws["A13"] = "Designated output"
    ws["B13"] = "Pale green fill — cell selected for reconstruction and reconciliation"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def build_workbook() -> Workbook:
    wb = Workbook()
    _build_guide(wb)
    _build_inputs(wb)
    _build_rollforward(wb)
    _build_controls(wb)
    return wb


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("Usage: python3 demo/build_case_4.py /path/to/output.xlsx", file=sys.stderr)
        return 2

    out_path = Path(argv[1])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) == 1:
        with tempfile.TemporaryDirectory() as tmp:
            default_path = Path(tmp) / "case_4_claims_reserve_roll_forward.xlsx"
            raise SystemExit(main([sys.argv[0], str(default_path)]))
    raise SystemExit(main(sys.argv))
