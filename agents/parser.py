"""Agent 1 — reads an .xlsx into a ParsedFile.

Pure Python, no LLM.

Every cell becomes one CellRecord holding BOTH its formula and its cached value.
That is not a convenience: Step 7 reconstructs from the formula and compares
against the cached value, so a record that could only hold one of them would
make the comparison impossible. The two are captured together or not at all.

Two dependency graphs are built, and they are not interchangeable.
`tab_dependency_graph` is a coarse overview. `cell_dependency_graph` is the one
that matters — it feeds circular reference detection and derivation chains. A
tab referencing another tab twice by two separate, non-circular cell chains is
ordinary; only the cell-level graph can tell that apart from a real cycle.
"""

import io
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime
from typing import Optional, Union

import networkx as nx
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from core.models import CellRecord, ParsedFile, WorkbookMeta
from core.workbook_identity import sha256_bytes

_TAB_REF_PATTERN = re.compile(r"'([^']+)'!|([A-Za-z_][A-Za-z0-9_.]*)!")
_NUMBER_LIKE_PATTERN = re.compile(r"^-?[\d,]+\.?\d*%?$")
_ERROR_VALUES = {"#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#N/A", "#NULL!", "#NUM!"}

# A cell reference, optionally sheet-qualified, optionally a range.
# The lookarounds keep function names out: "SUM(" is followed by "(", and
# "LOG10" would otherwise read as column LOG row 10.
_CELL_REF_PATTERN = re.compile(
    r"(?<![A-Za-z0-9_$!.])"
    r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"(\$?[A-Za-z]{1,3}\$?[0-9]{1,7})"
    r"(?::(\$?[A-Za-z]{1,3}\$?[0-9]{1,7}))?"
    r"(?![A-Za-z0-9_(])"
)

_STRING_LITERAL_PATTERN = re.compile(r'"[^"]*"')

# A range wider than this is recorded as an edge to its endpoints only, with a
# warning. Expanding C1:C1048576 cell by cell would hang the parse.
_MAX_RANGE_EXPANSION = 5000


def parse_workbook(workbook_bytes: bytes) -> ParsedFile:
    """Parse the exact bytes that were confirmed at Gate 1.

    Takes bytes rather than a path deliberately. A path is a mutable reference:
    this function reads the workbook five separate times (formula mode, data-only
    mode, metadata, VBA detection, and the hash), and against a path each of
    those is an independent window in which the file could change underneath us.
    The workbook a reviewer confirmed could then differ from the one parsed, with
    nothing to detect it.

    Bytes cannot change underneath the caller, so the invariant holds by
    construction: the bytes hashed for Gate 1 are the bytes parsed here. Each
    reader gets its OWN BytesIO over the same immutable sequence — sharing one
    buffer would rely on undocumented seek behaviour in openpyxl.
    """
    warnings: list[str] = []

    workbook_meta = _read_workbook_meta(workbook_bytes)

    wb_formulas = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    wb_values = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)

    raw_tab_names = list(wb_formulas.sheetnames)
    tab_names, dedupe_warnings = _dedupe_tab_keys(raw_tab_names)
    name_to_key = dict(zip(raw_tab_names, tab_names))
    warnings.extend(dedupe_warnings)
    warnings.extend(_check_sounds_like_duplicates(raw_tab_names))

    cells: dict[str, CellRecord] = {}
    external_links: list[str] = []
    named_ranges: dict[str, str] = {
        f"workbook::{name}": str(defn.value) for name, defn in wb_formulas.defined_names.items()
    }

    tab_graph = nx.DiGraph()
    tab_graph.add_nodes_from(tab_names)
    cell_graph = nx.DiGraph()

    for original_name, tab_key in zip(raw_tab_names, tab_names):
        ws_formulas = wb_formulas[original_name]
        ws_values = wb_values[original_name]

        for name, defn in ws_formulas.defined_names.items():
            named_ranges[f"{tab_key}::{name}"] = str(defn.value)

        for merged_range in ws_formulas.merged_cells.ranges:
            warnings.append(f"{tab_key}: merged cell range {merged_range}")

        max_row = ws_formulas.max_row or 0
        max_col = ws_formulas.max_column or 0
        if max_row == 0 or max_col == 0:
            # A blank tab is recorded by name with zero cells, not skipped and
            # not an error.
            continue

        for row in ws_formulas.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                raw_value = cell.value
                cached_value = ws_values[cell.coordinate].value

                if raw_value is None and cached_value is None:
                    continue

                key = f"{tab_key}!{cell.coordinate}"
                formula = _formula_text(raw_value, key, warnings)

                if formula is not None:
                    literal_value = None
                else:
                    literal_value = raw_value if raw_value is not None else cached_value
                    cached_value = literal_value

                record = _build_cell_record(
                    key=key,
                    formula=formula,
                    cached_value=cached_value,
                    number_format=cell.number_format,
                    calc_mode=workbook_meta.calc_mode,
                    warnings=warnings,
                )
                cells[key] = record
                cell_graph.add_node(key)

                if formula is None:
                    continue

                if "[" in formula and ".xls" in formula:
                    external_links.append(formula)

                for ref_name in _extract_tab_references(formula):
                    target_key = name_to_key.get(ref_name)
                    if target_key and target_key != tab_key:
                        tab_graph.add_edge(tab_key, target_key)

                for dependency in _extract_cell_references(
                    formula, own_tab=tab_key, name_to_key=name_to_key, key=key, warnings=warnings
                ):
                    cell_graph.add_edge(key, dependency)

    return ParsedFile(
        tab_names=tab_names,
        cells=cells,
        named_ranges=named_ranges,
        external_links=external_links,
        has_vba=_detect_vba(workbook_bytes),
        workbook_meta=workbook_meta,
        tab_dependency_graph=nx.to_dict_of_lists(tab_graph),
        cell_dependency_graph=nx.to_dict_of_lists(cell_graph),
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# cell records
# ---------------------------------------------------------------------------


def _formula_text(raw_value: object, key: str, warnings: list[str]) -> Optional[str]:
    """The formula string for a cell, or None if it holds a literal."""
    if isinstance(raw_value, ArrayFormula):
        warnings.append(f"{key}: array formula")
        return raw_value.text
    if isinstance(raw_value, str) and raw_value.startswith("{="):
        warnings.append(f"{key}: array formula")
        return raw_value
    if isinstance(raw_value, str) and raw_value.startswith("="):
        return raw_value
    return None


def _build_cell_record(
    key: str,
    formula: Optional[str],
    cached_value: object,
    number_format: str,
    calc_mode: str,
    warnings: list[str],
) -> CellRecord:
    data_type, is_error, error_type = _classify(cached_value, formula, key, warnings)

    # Stale means the cached value cannot be trusted as current — either it was
    # never computed, or the workbook doesn't recompute automatically.
    is_stale = formula is not None and (cached_value is None or calc_mode == "manual")
    if formula is not None and cached_value is None:
        warnings.append(f"{key}: formula has no cached value (never recalculated)")

    return CellRecord(
        cell_ref=key,
        formula=formula,
        cached_value=cached_value if isinstance(cached_value, (int, float, str, bool)) else _stringify(cached_value),
        data_type=data_type,
        number_format=number_format or "General",
        is_error=is_error,
        error_type=error_type,
        is_stale=is_stale,
    )


def _classify(
    value: object, formula: Optional[str], key: str, warnings: list[str]
) -> tuple[str, bool, Optional[str]]:
    """Data type from the value actually observed — never coerced or guessed.

    A number stored as text stays text, with a warning. Deciding it was "really"
    a number is exactly the kind of silent correction that hides a defect in the
    source workbook.
    """
    if isinstance(value, str) and value.strip() in _ERROR_VALUES:
        error = value.strip()
        warnings.append(f"{key}: formula error {error}")
        return "error", True, error
    if isinstance(value, bool):
        return "boolean", False, None
    if isinstance(value, (datetime, date)):
        return "date", False, None
    if isinstance(value, (int, float)):
        return "number", False, None
    if isinstance(value, str):
        if _looks_like_number(value):
            warnings.append(f"{key}: number stored as text ({value!r})")
        return "text", False, None
    if value is None:
        # A formula whose value was never computed has no observed type. The
        # model's vocabulary has no "unknown", so this reads as blank — is_stale
        # is what carries the real story, and a warning is recorded above.
        return "blank", False, None
    warnings.append(f"{key}: unexpected value type {type(value).__name__}, captured as text")
    return "text", False, None


def _stringify(value: object) -> Optional[str]:
    return None if value is None else str(value)


# ---------------------------------------------------------------------------
# workbook-level metadata
# ---------------------------------------------------------------------------


def _read_workbook_meta(workbook_bytes: bytes) -> WorkbookMeta:
    # The same canonical hash the UI showed and the human confirmed at Gate 1 —
    # computed by one helper so the two can never drift apart.
    workbook_hash = sha256_bytes(workbook_bytes)

    calc_mode = "unknown"
    fully_calculated_on_load = None
    app_version = None

    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
        names = archive.namelist()

        if "xl/workbook.xml" in names:
            root = ET.fromstring(archive.read("xl/workbook.xml"))
            calc_pr = next((el for el in root.iter() if el.tag.endswith("}calcPr") or el.tag == "calcPr"), None)
            if calc_pr is not None:
                raw_mode = calc_pr.get("calcMode")
                # Absent or unrecognised stays "unknown". Never assume automatic.
                if raw_mode == "manual":
                    calc_mode = "manual"
                elif raw_mode in ("auto", "autoNoTable"):
                    calc_mode = "automatic"
                raw_full = calc_pr.get("fullCalcOnLoad")
                if raw_full is not None:
                    fully_calculated_on_load = raw_full in ("1", "true")

        if "docProps/app.xml" in names:
            app_root = ET.fromstring(archive.read("docProps/app.xml"))
            application = _first_text(app_root, "Application")
            version = _first_text(app_root, "AppVersion")
            app_version = " ".join(part for part in (application, version) if part) or None

    return WorkbookMeta(
        calc_mode=calc_mode,
        workbook_hash=workbook_hash,
        app_version=app_version,
        fully_calculated_on_load=fully_calculated_on_load,
    )


def _first_text(root: ET.Element, local_name: str) -> Optional[str]:
    for element in root.iter():
        if element.tag.endswith("}" + local_name) or element.tag == local_name:
            return (element.text or "").strip() or None
    return None


# ---------------------------------------------------------------------------
# cell-level dependencies
# ---------------------------------------------------------------------------


def _extract_cell_references(
    formula: str, own_tab: str, name_to_key: dict, key: str, warnings: list[str]
) -> list[str]:
    """Every cell this formula directly references, as fully-qualified keys.

    Ranges are expanded cell by cell, because the graph exists to answer "which
    cells feed this one" — a range edge would answer a different question.
    """
    body = _STRING_LITERAL_PATTERN.sub('""', formula)
    references: list[str] = []

    for quoted_sheet, plain_sheet, start, end in _CELL_REF_PATTERN.findall(body):
        sheet_name = quoted_sheet or plain_sheet
        tab_key = name_to_key.get(sheet_name, sheet_name) if sheet_name else own_tab

        if not end:
            references.append(f"{tab_key}!{_normalize(start)}")
            continue

        expanded = _expand_range(_normalize(start), _normalize(end))
        if expanded is None:
            warnings.append(
                f"{key}: range {start}:{end} is too large to expand cell by cell; "
                f"only its endpoints are in the dependency graph"
            )
            references.extend([f"{tab_key}!{_normalize(start)}", f"{tab_key}!{_normalize(end)}"])
            continue
        references.extend(f"{tab_key}!{cell}" for cell in expanded)

    return references


def _normalize(ref: str) -> str:
    """A1-style reference with absolute markers removed and column upper-cased."""
    return ref.replace("$", "").upper()


def _expand_range(start: str, end: str) -> Optional[list[str]]:
    start_col, start_row = _split_ref(start)
    end_col, end_row = _split_ref(end)
    if None in (start_col, start_row, end_col, end_row):
        return [start, end]

    col_lo, col_hi = sorted((column_index_from_string(start_col), column_index_from_string(end_col)))
    row_lo, row_hi = sorted((start_row, end_row))
    if (col_hi - col_lo + 1) * (row_hi - row_lo + 1) > _MAX_RANGE_EXPANSION:
        return None

    return [
        f"{get_column_letter(col)}{row}"
        for col in range(col_lo, col_hi + 1)
        for row in range(row_lo, row_hi + 1)
    ]


def _split_ref(ref: str) -> tuple[Optional[str], Optional[int]]:
    match = re.fullmatch(r"([A-Z]{1,3})([0-9]{1,7})", ref)
    if not match:
        return None, None
    return match.group(1), int(match.group(2))


# ---------------------------------------------------------------------------
# tab names, external links, VBA
# ---------------------------------------------------------------------------


def _dedupe_tab_keys(raw_names: list[str]) -> tuple[list[str], list[str]]:
    keys: list[str] = []
    seen: dict[str, int] = {}
    warnings: list[str] = []
    for name in raw_names:
        if name not in seen:
            seen[name] = 0
            keys.append(name)
        else:
            seen[name] += 1
            new_key = f"{name}_{seen[name]}"
            warnings.append(
                f"duplicate tab name '{name}' renamed to '{new_key}' for cell-key purposes"
            )
            keys.append(new_key)
    return keys, warnings


def _check_sounds_like_duplicates(raw_names: list[str]) -> list[str]:
    warnings: list[str] = []
    normalized_seen: dict[str, str] = {}
    for name in raw_names:
        normalized = name.strip().lower()
        first_seen = normalized_seen.get(normalized)
        if first_seen is not None and first_seen != name:
            warnings.append(f"tab names '{first_seen}' and '{name}' look like duplicates")
        else:
            normalized_seen.setdefault(normalized, name)
    return warnings


def _extract_tab_references(formula: str) -> set[str]:
    refs = set()
    for quoted, unquoted in _TAB_REF_PATTERN.findall(formula):
        name = quoted or unquoted
        if name:
            refs.add(name)
    return refs


def _looks_like_number(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    return bool(_NUMBER_LIKE_PATTERN.match(stripped))


def _detect_vba(workbook_bytes: bytes) -> bool:
    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
        return "xl/vbaProject.bin" in archive.namelist()
